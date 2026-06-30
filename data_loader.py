"""
data_loader.py — Universal data ingestion for the mentorship matcher
=====================================================================
Loads mentors and students from either:
  A) Excel file (.xlsx)         — mentorship_data.xlsx
  B) SQLite database (.db)      — mentorship.db
  C) PostgreSQL / MySQL         — via connection string

The rest of the pipeline (v7 matcher) calls:
    mentors, students = load_data(source="excel", path="mentorship_data.xlsx")
    mentors, students = load_data(source="sqlite", path="mentorship.db")
    mentors, students = load_data(source="postgres", url="postgresql://...")

Returns two lists of dicts with identical structure regardless of source.
"""

import pandas as pd
import sqlite3
import os
from typing import Literal

# ─────────────────────────────────────────
# CONFIG — CGPA thresholds for tier logic
# ─────────────────────────────────────────

CGPA_SUPPORT_BELOW  = 6.0
CGPA_INTEREST_ABOVE = 8.0

# ─────────────────────────────────────────
# COLUMN SPECS
# List columns that are comma-separated
# strings and need to be parsed into lists
# ─────────────────────────────────────────

MENTOR_LIST_COLS  = ["skills", "can_help_with", "specialises_in", "industries"]
STUDENT_LIST_COLS = ["goals", "issues", "interests", "industries"]


# ─────────────────────────────────────────
# PARSING HELPERS
# ─────────────────────────────────────────

def _parse_list(value) -> list:
    """'python, ml, sql'  →  ['python', 'ml', 'sql']"""
    if pd.isna(value) or value == "":
        return []
    return [v.strip() for v in str(value).split(",") if v.strip()]

def _assign_tier(cgpa: float) -> str:
    if cgpa < CGPA_SUPPORT_BELOW:
        return "support"
    elif cgpa >= CGPA_INTEREST_ABOVE:
        return "interest"
    return "standard"

def _normalise_mentors(df: pd.DataFrame) -> list:
    """Convert a raw mentor DataFrame into the list-of-dicts the matcher expects."""
    required = ["mentor_id", "name", "bio", "skills", "can_help_with",
                "specialises_in", "industries", "availability", "track"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Mentor data missing columns: {missing}")

    mentors = []
    for _, row in df.iterrows():
        m = {
            "id":             str(row["mentor_id"]),
            "name":           str(row["name"]),
            "bio":            str(row["bio"]),
            "skills":         _parse_list(row["skills"]),
            "can_help_with":  _parse_list(row["can_help_with"]),
            "specialises_in": _parse_list(row["specialises_in"]),
            "industries":     _parse_list(row["industries"]),
            "availability":   str(row["availability"]).strip().lower(),
            "track":          str(row["track"]).strip().lower(),
        }
        mentors.append(m)
    return mentors

def _normalise_students(df: pd.DataFrame) -> list:
    """Convert a raw student DataFrame into the list-of-dicts the matcher expects."""
    required = ["student_id", "name", "bio", "cgpa", "goals", "issues",
                "interests", "industries", "availability", "track"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Student data missing columns: {missing}")

    students = []
    for _, row in df.iterrows():
        cgpa = float(row["cgpa"])
        s = {
            "id":           str(row["student_id"]),
            "name":         str(row["name"]),
            "bio":          str(row["bio"]),
            "cgpa":         cgpa,
            "tier":         _assign_tier(cgpa),
            "goals":        _parse_list(row["goals"]),
            "issues":       _parse_list(row["issues"]),
            "interests":    _parse_list(row["interests"]),
            "industries":   _parse_list(row["industries"]),
            "availability": str(row["availability"]).strip().lower(),
            "track":        str(row["track"]).strip().lower(),
        }
        students.append(s)
    return students


# ─────────────────────────────────────────
# LOADERS — one per source type
# ─────────────────────────────────────────

def _load_excel(path: str):
    if not os.path.exists(path):
        raise FileNotFoundError(f"Excel file not found: {path}")
    xls = pd.ExcelFile(path)
    if "Mentors" not in xls.sheet_names or "Students" not in xls.sheet_names:
        raise ValueError(f"Excel file must have 'Mentors' and 'Students' sheets. "
                         f"Found: {xls.sheet_names}")
    mentors_df  = pd.read_excel(path, sheet_name="Mentors")
    students_df = pd.read_excel(path, sheet_name="Students")
    return mentors_df, students_df


def _load_sqlite(path: str):
    if not os.path.exists(path):
        raise FileNotFoundError(f"SQLite database not found: {path}")
    conn = sqlite3.connect(path)
    try:
        mentors_df  = pd.read_sql("SELECT * FROM mentors",  conn)
        students_df = pd.read_sql("SELECT * FROM students", conn)
    finally:
        conn.close()
    return mentors_df, students_df


def _load_postgres(url: str):
    try:
        from sqlalchemy import create_engine
    except ImportError:
        raise ImportError("Install sqlalchemy and psycopg2: pip install sqlalchemy psycopg2-binary")
    engine = create_engine(url)
    with engine.connect() as conn:
        mentors_df  = pd.read_sql("SELECT * FROM mentors",  conn)
        students_df = pd.read_sql("SELECT * FROM students", conn)
    return mentors_df, students_df


def _load_mysql(url: str):
    try:
        from sqlalchemy import create_engine
    except ImportError:
        raise ImportError("Install sqlalchemy and pymysql: pip install sqlalchemy pymysql")
    engine = create_engine(url)
    with engine.connect() as conn:
        mentors_df  = pd.read_sql("SELECT * FROM mentors",  conn)
        students_df = pd.read_sql("SELECT * FROM students", conn)
    return mentors_df, students_df


# ─────────────────────────────────────────
# PUBLIC API — single entry point
# ─────────────────────────────────────────

def load_data(
    source: Literal["excel", "sqlite", "postgres", "mysql"] = "excel",
    path:   str = "mentorship_data.xlsx",
    url:    str = None,
) -> tuple[list, list]:
    """
    Load mentors and students from the specified source.

    Parameters
    ----------
    source  : "excel" | "sqlite" | "postgres" | "mysql"
    path    : file path (for excel / sqlite)
    url     : connection string (for postgres / mysql)
              postgres → "postgresql://user:pass@host:5432/dbname"
              mysql    → "mysql+pymysql://user:pass@host:3306/dbname"

    Returns
    -------
    mentors  : list of dicts
    students : list of dicts
    """
    print(f"Loading data from {source.upper()}...")

    if source == "excel":
        mentors_df, students_df = _load_excel(path)
    elif source == "sqlite":
        mentors_df, students_df = _load_sqlite(path)
    elif source == "postgres":
        if not url:
            raise ValueError("Provide url= for postgres source.")
        mentors_df, students_df = _load_postgres(url)
    elif source == "mysql":
        if not url:
            raise ValueError("Provide url= for mysql source.")
        mentors_df, students_df = _load_mysql(url)
    else:
        raise ValueError(f"Unknown source '{source}'. Use excel/sqlite/postgres/mysql.")

    mentors  = _normalise_mentors(mentors_df)
    students = _normalise_students(students_df)

    # Print tier breakdown
    tiers = {"support": 0, "standard": 0, "interest": 0}
    for s in students:
        tiers[s["tier"]] += 1

    print(f"  Loaded {len(mentors)} mentors, {len(students)} students")
    print(f"  Tiers → support: {tiers['support']}  "
          f"standard: {tiers['standard']}  interest: {tiers['interest']}")
    return mentors, students


# ─────────────────────────────────────────
# EXPORT HELPER — save assignments back out
# ─────────────────────────────────────────

def export_assignments(assignments: dict, students: list, mentors: list,
                       score_matrix, out_path: str = "assignments.xlsx",
                       unassigned: list = None,
                       unassigned_reasons: dict = None):
    """
    Write the final assignments to an Excel file.

    Parameters
    ----------
    assignments  : {student_idx: mentor_idx}
    students     : list of student dicts
    mentors      : list of mentor dicts
    score_matrix : numpy array (students × mentors), normalised 1–10
    out_path     : where to write the output
    """
    unassigned         = unassigned or []
    unassigned_reasons = unassigned_reasons or {}
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Assignments sheet ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Assignments"

    TIER_COLORS = {"support": "FFD7D7", "standard": "FFFBCC", "interest": "D7F5DC"}
    HEADER_FILL = PatternFill("solid", start_color="1F4E79")
    HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    thin        = Side(style="thin", color="BBBBBB")
    BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Student ID", "Student Name", "CGPA", "Tier",
               "Mentor ID", "Mentor Name", "Match Score", "Shared Skills"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill = HEADER_FONT, HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    ws.row_dimensions[1].height = 28

    for row_idx, (si, mi) in enumerate(sorted(assignments.items()), 2):
        s, m  = students[si], mentors[mi]
        score = round(float(score_matrix[si][mi]), 2)
        shared = ", ".join(set(m["skills"]) & set(s["goals"])) or "—"
        tier_fill = PatternFill("solid", start_color=TIER_COLORS[s["tier"]])

        values = [s["id"], s["name"], s["cgpa"], s["tier"].upper(),
                  m["id"], m["name"], score, shared]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.fill   = tier_fill
            c.border = BORDER
            c.font   = Font(name="Arial", size=10)
            c.alignment = Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{len(assignments)+1}"
    for col, w in enumerate([10,16,6,10,10,16,10,28], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── Summary sheet ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    roster_sizes   = {}
    roster_scores  = {}
    for si, mi in assignments.items():
        roster_sizes[mi]  = roster_sizes.get(mi, 0) + 1
        roster_scores.setdefault(mi, []).append(float(score_matrix[si][mi]))

    sum_headers = ["Mentor ID", "Mentor Name", "Track",
                   "Students Assigned", "Avg Score", "Min Score"]
    for col, h in enumerate(sum_headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font, c.fill = HEADER_FONT, HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER

    for row_idx, (mi, mentor) in enumerate(enumerate(mentors), 2):
        scores = roster_scores.get(mi, [])
        values = [mentor["id"], mentor["name"], mentor["track"],
                  roster_sizes.get(mi, 0),
                  round(sum(scores)/len(scores), 2) if scores else 0,
                  round(min(scores), 2) if scores else 0]
        for col, val in enumerate(values, 1):
            c = ws2.cell(row=row_idx, column=col, value=val)
            c.border = BORDER
            c.font   = Font(name="Arial", size=10)
            c.alignment = Alignment(horizontal="left", vertical="center")

    for col, w in enumerate([10,16,14,16,10,10], 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    wb.save(out_path)
    print(f"\nAssignments exported to: {out_path}")

    ws3 = wb.create_sheet("Unassigned")
    UNASSIGNED_FILL = PatternFill("solid", start_color="F2F2F2")  # grey
 
    ua_headers = ["Student ID", "Student Name", "CGPA", "Tier",
                  "Best Mentor (if capacity existed)", "Best Score", "Reason"]
    for col, h in enumerate(ua_headers, 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font, c.fill = HEADER_FONT, HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    ws3.row_dimensions[1].height = 28
 
    if unassigned:
        import numpy as np
        for row_idx, si in enumerate(unassigned, 2):
            s        = students[si]
            best_j   = int(np.argmax(score_matrix[si]))
            best_sc  = round(float(score_matrix[si][best_j]), 2)
            best_m   = mentors[best_j]["name"]
            reason   = unassigned_reasons.get(si, "All mentor capacities full.")
 
            values = [s["id"], s["name"], s["cgpa"], s["tier"].upper(),
                      best_m, best_sc, reason]
            for col, val in enumerate(values, 1):
                c = ws3.cell(row=row_idx, column=col, value=val)
                c.fill      = UNASSIGNED_FILL
                c.border    = BORDER
                c.font      = Font(name="Arial", size=10,
                                   color="888888")  # greyed-out text
                c.alignment = Alignment(horizontal="left", vertical="center",
                                        wrap_text=True)
    else:
        # No unassigned students — note it
        c = ws3.cell(row=2, column=1, value="All students were successfully assigned.")
        c.font = Font(name="Arial", size=10, italic=True, color="555555")
 
    ws3.freeze_panes = "A2"
    if unassigned:
        ws3.auto_filter.ref = f"A1:G{len(unassigned)+1}"
    for col, w in enumerate([10, 16, 6, 10, 18, 10, 52], 1):
        ws3.column_dimensions[get_column_letter(col)].width = w
 
    wb.save(out_path)
    print(f"\nAssignments exported to: {out_path}")
    print(f"  Sheets: Assignments ({len(assignments)} rows)  |  "
          f"Unassigned ({len(unassigned)} rows)  |  Summary")
 

# ─────────────────────────────────────────
# SQLITE SETUP HELPER
# Creates a fresh database from the Excel file
# ─────────────────────────────────────────

def excel_to_sqlite(excel_path: str, db_path: str):
    """
    One-time migration: copy Excel data into a SQLite database.
    After this, you can use source='sqlite' going forward.
    """
    mentors_df, students_df = _load_excel(excel_path)
    conn = sqlite3.connect(db_path)
    mentors_df.to_sql("mentors",  conn, if_exists="replace", index=False)
    students_df.to_sql("students", conn, if_exists="replace", index=False)
    conn.close()
    print(f"Created SQLite database: {db_path}")
    print(f"  Tables: mentors ({len(mentors_df)} rows), "
          f"students ({len(students_df)} rows)")


# ─────────────────────────────────────────
# QUICK TEST — run this file directly to
# verify loading works on your data file
# ─────────────────────────────────────────

if __name__ == "__main__":
    import sys

    # ── Test Excel loading ──────────────────
    print("=" * 50)
    print("TEST 1: Load from Excel")
    print("=" * 50)
    mentors, students = load_data(source="excel", path="mentorship_data.xlsx")
    print(f"\nFirst mentor : {mentors[0]['name']}  |  skills: {mentors[0]['skills']}")
    print(f"First student: {students[0]['name']}  |  tier: {students[0]['tier']}  "
          f"cgpa: {students[0]['cgpa']}")

    # ── Create SQLite from Excel ─────────────
    print("\n" + "=" * 50)
    print("TEST 2: Convert Excel → SQLite")
    print("=" * 50)
    excel_to_sqlite("mentorship_data.xlsx", "mentorship.db")

    # ── Test SQLite loading ─────────────────
    print("\n" + "=" * 50)
    print("TEST 3: Load from SQLite")
    print("=" * 50)
    mentors2, students2 = load_data(source="sqlite", path="mentorship.db")
    assert len(mentors2) == len(mentors), "Mentor count mismatch!"
    assert len(students2) == len(students), "Student count mismatch!"
    print("SQLite load matches Excel load ✓")
